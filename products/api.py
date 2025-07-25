from ninja import  Router, Query

from ninja.files import UploadedFile


# router.py
from .models import ProductListingImage, Category, FeatureGroup, FeatureTemplate, Product, ProductListing, Feature, ReturnExchangePolicy, Variant
from .schemas import ( 
    CategoryCreateSchema, CategoryOutSchema, CategoryUpdateSchema,
    # CategoryFeatureValuesOutSchema,
    CategorySchema, CategoryParentChildrenOutSchema,
    FeatureGroupOutSchema,
    FeatureTemplateOutSchema,
    ProductCreateSchema, ProductOutSchema, ProductOutOneSchema, ProductUpdateSchema,
    ProductListingUpdateSchema, ProductListingCreateSchema, ProductListingOutSchema, ProductListingOneOutSchema,
    FeatureOutSchema,
    ProductListingImageOutSchema, VariantSchema, VariantCreateSchema,VariantUpdateSchema,
    ReturnExchangePolicySchema, ReturnExchangePolicyCreateSchema, ReturnExchangePolicyUpdateSchema
)
from django.shortcuts import get_object_or_404
from utils.pagination import PaginatedResponseSchema, paginate_queryset
from ninja.files import UploadedFile

import json 
from django.db.models import Min, Max, Count

from utils.cache import cache_response


from django.db.models import Q

from ninja_jwt.authentication import JWTAuth

from typing import Optional

from ninja import File, Form

from ninja import Schema
router = Router()
class ErrorSchema(Schema):
    detail: str

from openpyxl import load_workbook
from users.models import Entity
from users.schemas import EntityOut2Schema
from taxations.models import TaxCategory
from estores.models import EStore
from ast import literal_eval
from django.db import IntegrityError

############################ Product Listing Upload from files ############################

def parse_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['true', '1', 'yes']
    if isinstance(value, int):
        return value == 1
    return False


def get_or_create_entity(name, entity_type, estore_id=None):
    if not name:
        return None
    
    if estore_id:
        estore=EStore.objects.filter(id=estore_id).first()
        return Entity.objects.get_or_create(name=name.strip(), entity_type=entity_type, estore=estore)[0]
    return Entity.objects.get_or_create(name=name.strip(), entity_type=entity_type)[0]

def get_or_create_category(name):
    if not name or str(name).strip() == "":
        return None
    return Category.objects.get_or_create(name=name.strip())[0]

def get_category_by_id(category_id):
    if not category_id:
        return None
    return Category.objects.filter(id=category_id).first()


def get_tax_category_by_id(tax_id):
    if not tax_id:
        return None
    return TaxCategory.objects.filter(id=tax_id).first()


@router.post("/upload-products/")
def upload_products_from_excel(request, 
        seller_id: int = None,
        estore_id: int = None,
        file: UploadedFile = File(...)
    ):

    wb = load_workbook(filename=file.file)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    created_count = 0
    errors = []

    empty_row_count = 0

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        
        if not any(cell not in (None, "", " ") for cell in row):
            empty_row_count += 1
            if empty_row_count >= 2:
                break  # stop after 2 consecutive empty rows
            continue  # skip this empty row
        else:
            empty_row_count = 0  # reset if row is not empty


        row_data = dict(zip(headers, row))
        try:
            product, _ = Product.objects.get_or_create(
                name=row_data['product_name'],
                defaults={
                    'about': row_data.get('product_about') or "",
                    'description': row_data.get('product_description') or "",
                    'base_price': row_data.get('base_price') or 0,
                    'is_service': parse_bool(row_data.get('is_service')),
                    'category': get_category_by_id(row_data.get('category_id')),
                    'brand': get_or_create_entity(row_data.get('brand_name'), 'brand', row_data.get("estore_id")),
                    'unit_size': row_data.get('unit_size') or 1,
                    'size_unit': row_data.get('size_unit') or "",
                    'tax_category': get_tax_category_by_id(row_data.get('tax_category_id')),
                }
            )

            # Create Variant
            try:
                attributes = literal_eval(row_data.get('variant_attributes') or "[]")
            except:
                attributes = []
            
            if row_data.get('variant_name'):
                variant, _ = Variant.objects.get_or_create(
                    product=product,
                    name=row_data['variant_name'],
                    defaults={'attributes': attributes}
                )
            else:
                variant = None

            # Create ProductListing
            ProductListing.objects.create(
                product=product,
                variant=variant,
                name="",
                price=row_data.get('price') or 0,
                mrp=row_data.get('mrp') or 0,
                stock=row_data.get('stock') or 0,
                approved=False,
                featured=False,
                units_per_pack=row_data.get('units_per_pack') or 1,
                seller=Entity.objects.filter(id=row_data.get('seller_id')).first(),
                estore=EStore.objects.filter(id=row_data.get('estore_id')).first()
            )

            created_count += 1

        except Exception as e:
            errors.append({
                "row": idx,
                "error": str(e),
                "data": row_data
            })

    return {
        "status": "success" if not errors else "partial",
        "created": created_count,
        "errors": errors
    }

############################ Category ############################
# Create Category

@router.post("/categories/", response=CategoryOutSchema)
def create_category(request, payload: CategoryCreateSchema):
    # category = Category(**payload.dict())
    data = payload

    # print(data);

    if data.parent_id:
        # Adding as a child category
        parent = get_object_or_404(Category, id=data.parent_id)
        category = parent.add_child(
            name=data.name,
            description=data.description,
            feature_names=data.feature_names,
        )
    else:
        # Adding as a root category
        category = Category.add_root(
            name=data.name,
            description=data.description,
            feature_names=data.feature_names,
        )
    
    category.save()
    return category

# Read Users (List)
@router.get("/categories/", response=PaginatedResponseSchema)
@cache_response()
def categories(
        request,
        page: int = Query(1, description="Page number"),
        page_size: int = Query(10, description="Number of items per page"),
        estore_id: Optional[int] = Query(None, description="Filter by EStore ID"),
        level: Optional[int] = Query(None, description="Filter by category level"),
        has_blogs: Optional[bool] = Query(None, description="Filter categories that have associated blogs"),
        category_type: Optional[str] = Query("product", description="Type of category"),
        search: Optional[str] = Query(None, description="Search by category name"),
    ):
    qs = Category.objects.filter(approved=True)

    query = ""

    if search:
        qs = qs.filter(name__icontains=search)
        query = query + "&search=" + search

    if estore_id is not None:
        qs = qs.filter(estore__id=estore_id)
        query = query + "&estore_id=" + str(estore_id)

    if level:
        qs = qs.filter(level=level)
        query = query + "&level=" + str(level)

    if category_type:
        qs = qs.filter(category_type=category_type)
        query = query + "&category_type=" + category_type

    if has_blogs:
        qs = qs.filter(category_blogs__isnull=False).distinct()
        query = query + "&has_blogs=" + str(has_blogs)

    return paginate_queryset(request, qs, CategoryOutSchema, page, page_size)

@router.get("/categories/parents-children/{category_id}/", response=CategoryParentChildrenOutSchema)
@cache_response()
def retrieve_category_parents_children(request, category_id: int):
    category = get_object_or_404(Category, id=category_id)

    def get_all_parents(category):
        parents = []
        while category.get_parent():
            category = category.get_parent()
            parents.insert(0,category)
        return parents
    
    parents = get_all_parents(category)

    # Get the children categories
    children = category.get_children()

    return {
        "parents": [CategorySchema.from_orm(parent) for parent in parents],
        "children": [CategorySchema.from_orm(child) for child in children],
    }


@router.get("/categories/siblings/{category_id}/", response=list[CategoryOutSchema])
@cache_response()
def retrieve_category_siblings(request, category_id: int):
    category = get_object_or_404(Category, id=category_id)
    
    # Get the parent of the current category
    parent = category.get_parent()
    
    # If category has a parent, get its siblings (excluding itself)
    if parent:
        siblings = parent.get_children()
    else:
        # If no parent, get all root categories (excluding itself)
        siblings = Category.get_root_nodes()
    
    return [CategoryOutSchema.from_orm(sibling) for sibling in siblings]


# Read Single User (Retrieve)
@router.get("/categories/{category_id}/", response=CategoryOutSchema)
def retrieve_category(request, category_id: int):
    category = get_object_or_404(Category, id=category_id)
    return category


@router.get("/categories/slug/{category_slug}/", response=CategoryOutSchema)
@cache_response()
def retrieve_category_slug(request, category_slug: str):
    category = get_object_or_404(Category, slug=category_slug)
    return CategoryOutSchema.from_orm(category)

# Update User
@router.put("/categories/{category_id}/", response=CategoryOutSchema)
def update_category(request, category_id: int, payload: CategoryUpdateSchema):
    category = get_object_or_404(Category, id=category_id)
    for attr, value in payload.dict().items():
        if value is not None:
            setattr(category, attr, value)
    category.save()
    return category

# Delete User
@router.delete("/categories/{category_id}/")
def delete_category(request, category_id: int):
    category = get_object_or_404(Category, id=category_id)
    category.delete()
    return {"success": True}


@router.get("/feature-groups/", response=PaginatedResponseSchema)
def featuregroups(request,  page: int = 1, page_size: int = 10, category_id:str = None , ordering: str = None,):
    qs = FeatureGroup.objects.all()

    if category_id:
        qs = qs.filter(category__id=category_id)

    if ordering:
        qs = qs.order_by(ordering)

    return paginate_queryset(request, qs, FeatureGroupOutSchema , page, page_size)



@router.get("/feature-templates/", response=PaginatedResponseSchema)
def featuretemplates(request,  page: int = 1, page_size: int = 10, feature_group_id:str = None , ordering: str = None,):
    qs = FeatureTemplate.objects.all()


    if feature_group_id:
        qs = qs.filter(feature_group__id=feature_group_id)

    if ordering:
        qs = qs.order_by(ordering)

    return paginate_queryset(request, qs, FeatureTemplateOutSchema , page, page_size)

############################ Product ############################
@router.post("/products/", response=ProductOutSchema,  auth=JWTAuth())
def create_product(request, payload: ProductCreateSchema):
    # locality = get_object_or_404(Locality, id=payload.locality_id)
    product = Product(**payload.dict())
    product.save()
    return product

# Read Products (List)
@router.get("/products/", response=PaginatedResponseSchema)
@cache_response()
def products(request,  page: int = 1, page_size: int = 10, category_id:str = None , ordering: str = None, seller_id: int = None):
    qs = Product.objects.all()

    query = ""

    if category_id:
        qs = qs.filter(category__id=category_id)
        query = query + "&category_id=" + category_id

    if seller_id:
        qs = qs.filter(product_listings__seller__id=seller_id).distinct()
        query = query + "&seller_id=" + str(seller_id)


    if ordering:
        qs = qs.order_by(ordering)
        query = query + "&ordering=" + ordering

    return paginate_queryset(request, qs, ProductOutSchema, page, page_size, query)

# Read Single Product (Retrieve)
@router.get("/products/{product_id}/", response=ProductOutOneSchema)
@cache_response()
def retrieve_product(request, product_id: int):

    product = get_object_or_404(Product.objects.prefetch_related('product_variants'), id=product_id)


    return {
        'id': product.id,
        'name': product.name,
        'about': product.about,
        'description': product.description,
        'size_unit': product.size_unit,
        'unit_size': product.unit_size,
        'important_info': product.important_info,
        'brand': EntityOut2Schema.from_orm(product.brand),
        'category': CategoryOutSchema.from_orm(product.category),
        'base_price': product.base_price,
        'is_service': product.is_service,
        'tax_category': product.tax_category if product.tax_category else None,
        'country_of_origin': product.country_of_origin,
        'created': product.created,
        'updated': product.updated,
        'variants': [VariantSchema.from_orm(v).model_dump() for v in product.product_variants.all()]
    }

# Update Product
@router.put("/products/{product_id}/", response=ProductOutSchema)
def update_product(request, product_id: int, payload: ProductUpdateSchema):
    product = get_object_or_404(Product, id=product_id)
    for attr, value in payload.dict().items():
        if value is not None:
            setattr(product, attr, value)
    product.save()
    return product

# Delete Product
@router.delete("/products/{product_id}/")
def delete_product(request, product_id: int):
    product = get_object_or_404(Product, id=product_id)
    product.delete()
    return {"success": True}


############################ Product Listing ####################


@router.post("/product-listings/", response=ProductListingOutSchema)
def create_product_listing(
    request,
    product_id: int = Form(...),
    name: str = Form(None),
    category_id: int = Form(None),
    brand_id: int = Form(None),
    manufacturer_id: int = Form(None),
    tax_category_id: int = Form(None),
    return_exchange_policy_id: int = Form(None),
    estore_id: int = Form(None),
    box_items: str = Form(None),
    features: str = Form(None),  # JSON string
    approved: bool = Form(False),
    featured: bool = Form(False),
    variant_id: int = Form(None),
    seller_id: int = Form(None),
    packer_id: int = Form(None),
    importer_id: int = Form(None),
    price: float = Form(...),
    mrp: float = Form(None),
    stock: int = Form(0),
    buy_limit: int = Form(10),
    rating: float = Form(5.0),
    review_count: int = Form(1),
    popularity: int = Form(100),
    main_image: UploadedFile = File(None)
):
    product_listing = ProductListing(
        product_id=product_id,
        name=name,
        category_id=category_id,
        brand_id=brand_id,
        manufacturer_id=manufacturer_id,
        tax_category_id=tax_category_id,
        return_exchange_policy_id=return_exchange_policy_id,
        estore_id=estore_id,
        box_items=box_items,
        approved=approved,
        featured=featured,
        variant_id=variant_id,
        seller_id=seller_id,
        packer_id=packer_id,
        importer_id=importer_id,
        price=price,
        mrp=mrp,
        stock=stock,
        buy_limit=buy_limit,
        rating=rating,
        review_count=review_count,
        popularity=popularity,
    )
    if features is not None:
        try:
            product_listing.features = json.loads(features)
        except Exception:
            product_listing.features = None
    if main_image is not None:
        if hasattr(main_image, 'size') and main_image.size > 0:
            product_listing.main_image = main_image
    product_listing.save()
    return product_listing

# Read ProductListings (List)
# /api/product_listings/?category_id=5&brand_ids=1&min_price=2000&feature_filters={"ram":["8GB"]}



@router.get("/product-listings/", response=PaginatedResponseSchema)
@cache_response()
def product_listings(
    request,
    page: int = 1,
    page_size: int = 10,
    category_id: str = None,
    seller_id: int = None,
    product_id:int = None,
    is_service: bool = None,
    search: str = None,
    ordering: str = None,
    featured: bool = None,
    brand_ids: str = Query(None, description="Comma-separated brand IDs"),
    min_price: int = Query(None, description="Minimum price"),
    max_price: int = Query(None, description="Maximum price"),
    feature_filters: str = Query(None, description="Feature filters as JSON string"),
    estore_id: int = Query(None, description="EStore ID"),
    approved: bool = Query(None, description="Approved"),
):
    qs = ProductListing.objects.all()
    query = ""

    # print("Product",category_id, brand_ids, min_price, max_price, feature_filters)


    if seller_id:
        qs = qs.filter(seller__id=seller_id)
        query = query + "&seller_id=" + str(seller_id)

    if product_id:
        qs = qs.filter(product__id=product_id)
        query = query + "&product_id=" + str(product_id)

        
    # Filter by category and its children
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
            descendants = category.get_descendants()
            qs = qs.filter(category__in=[category] + list(descendants))
            query = query + "&category_id=" + category_id
        except Category.DoesNotExist:
            return {"error": "Category not found"}
        
    if featured:
        qs = qs.filter(featured =featured)
        query = query + "&featured=" + str(featured)

    if estore_id:
        qs = qs.filter(estore__id=estore_id)
        query = query + "&estore_id=" + str(estore_id)

    if approved is not None:
        qs = qs.filter(approved=approved)
        query = query + "&approved=" + str(approved)

        # Filter by search term
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(product__name__icontains=search))
        query = query + "&search=" + search

    # Filter by brands
    if brand_ids:
        brand_id_list = brand_ids.split(",")
        qs = qs.filter(brand__id__in=brand_id_list)
        query = query + "&brand_ids=" + brand_ids

    # Filter by price range
    if min_price is not None:
        qs = qs.filter(price__gte=min_price)
        query = query + "&min_price=" + str(min_price)

    if max_price is not None:
        qs = qs.filter(price__lte=max_price)
        query = query + "&max_price=" + str(max_price)

    
    if is_service is not None:
        qs = qs.filter(is_service=is_service)
        query = query + "&is_service=" + str(is_service)
        
    # Filter by features
    if feature_filters:
        query = query + "&feature_filters=" + feature_filters
        try:
            feature_filters = json.loads(feature_filters)  # Parse the JSON string
            for feature_template_id, values in feature_filters.items():
                qs = qs.filter(
                    product_listing_features__feature_template__id=feature_template_id,
                    product_listing_features__value__in=values,
                )
        except Exception as e:
            return {"error": f"Feature filter error: {str(e)}"}

    # Ordering
    if ordering:
        query = query + "&ordering=" + ordering
        qs = qs.order_by(ordering)

    # Paginate the results
    return paginate_queryset(request, qs, ProductListingOutSchema, page, page_size, query)

# Update ProductListing (JSON, no file upload)
@router.put("/product-listings/{product_listing_id}/", response={200: ProductListingOutSchema, 400: ErrorSchema})
def update_product_listing(
    request,
    product_listing_id: int,
    payload: ProductListingUpdateSchema
):
    print("Received JSON payload:", payload.dict())
    product_listing = get_object_or_404(ProductListing, id=product_listing_id)
    # Check for variant_id uniqueness
    if payload.variant_id is not None:
        existing = ProductListing.objects.filter(variant_id=payload.variant_id).exclude(id=product_listing_id).first()
        if existing:
            return 400, {"detail": "This variant is already assigned to another listing."}
    for attr, value in payload.dict(exclude_unset=True).items():
        if value is not None:
            setattr(product_listing, attr, value)
    try:
        product_listing.save()
    except IntegrityError as e:
        return 400, {"detail": f"Database integrity error: {str(e)}"}
    return product_listing

# Update ProductListing main_image only (multipart/form-data)
@router.post("/product-listings/{product_listing_id}/update-image/", response=ProductListingOutSchema)
def update_product_listing_image(
    request,
    product_listing_id: int,
    main_image: UploadedFile = File(...)
):
    product_listing = get_object_or_404(ProductListing, id=product_listing_id)
    if main_image is not None:
        product_listing.main_image = main_image
        product_listing.save()
    return product_listing


@router.get("/sidebar-filters/", tags=["Sidebar filters"])
@cache_response()
def get_sidebar_filters(
    request, 
    category_id: str = None,
    search: str = None,
    is_service: bool = None,
    brand_ids: str = Query(None, description="Comma-separated brand IDs"),  # Example: '1,2,3'
    min_price: float = Query(None, description="Minimum price"),
    max_price: float = Query(None, description="Maximum price"),
    feature_filters: str = Query(None, description="Feature filters as JSON string"),  # Example: '{"1": ["4GB", "6GB"], "2": ["128GB"]}'
    estore_id: int = Query(None, description="EStore ID"),
    approved: bool = Query(None, description="Approved"),
    ):
    """
    API to fetch sidebar filters for product listings.
    """
    filters = {}

    # print(category_id, brand_ids, min_price, max_price, feature_filters)
    
    # Filter listings by category if category_id is provided
    qs = ProductListing.objects.all()
    
    # Filter by category
    
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
            descendants = category.get_descendants()
            qs = qs.filter(category__in=[category] + list(descendants))
        except Category.DoesNotExist:
            return {"error": "Category not found"}
        
    if estore_id:
        qs = qs.filter(estore__id=estore_id)
        query = query + "&estore_id=" + str(estore_id)

    if approved is not None:
        qs = qs.filter(approved=approved)
    
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(product__name__icontains=search))


    # Filter by brands
    if brand_ids:
        brand_id_list = brand_ids.split(",")  # Split comma-separated string into list
        qs = qs.filter(brand__id__in=brand_id_list)

    # Filter by price range
    if min_price is not None:
        qs = qs.filter(price__gte=min_price)
    if max_price is not None:
        qs = qs.filter(price__lte=max_price)

    if is_service is not None:
        qs = qs.filter(is_service=is_service)

    if feature_filters:
        try:
            for feature_template_id, values in feature_filters.items():
                qs = qs.filter(
                    product_listing_features__feature_template__id=feature_template_id,
                    product_listing_features__value__in=values,
                )
        except Exception as e:
            return {"error": f"Feature filter error: {str(e)}"}

    # Get brand filters
    filters['brands'] = list(
        qs.values('brand__id', 'brand__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # filters['categories'] = list(
    #     qs.values('category__id', 'category__name')
    #     .annotate(count=Count('id'))
    #     .order_by('-count')
    # )

    filters['features'] = list(
        qs.values('product_listing_features__feature_template__id', 'product_listing_features__feature_template__name','product_listing_features__value')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # print(filters)
    
    # Get price range
    price_range = qs.aggregate(
        min_price=Min('price'),
        max_price=Max('price')
    )

    filters['price_range'] = {
        "min_price": price_range.get('min_price', 0),
        "max_price": price_range.get('max_price', 0)
    }

    return filters

@router.get("/product-listings/related/{product_listing_id}/", response=PaginatedResponseSchema)
@cache_response()
def get_related_products(
    request,
    product_listing_id: int,
    page: int = 1,
    page_size: int = 10,
    ):
    """
    Get related products for a specific product listing based on category, brand, features, and price similarity.
    """
    # Get the original product listing
    product_listing = get_object_or_404(ProductListing, id=product_listing_id, approved=True)
    qs = ProductListing.objects.filter(approved=True).exclude(id=product_listing_id)

    # Base filters for related products
    related_filters = Q()

    # 1. Same category filter
    if product_listing.category:
        related_filters &= Q(category=product_listing.category)

    # 2. Same brand filter
    # if product_listing.brand:
    #     related_filters &= Q(brand=product_listing.brand)

    # 3. Similar price range (±50% of original price)
    price = float(product_listing.price)
    min_price = price * 0.5  # 50% below
    max_price = price * 1.5  # 50% above
    related_filters &= Q(price__gte=min_price) & Q(price__lte=max_price)

    # 4. Feature similarity
    # if product_listing.features:
    #     try:
    #         features = product_listing.features
    #         # Look for products with at least one matching feature
    #         feature_filters = Q()
    #         for category in features:
    #             for feature in features[category]:
    #                 feature_filters |= Q(features__contains={category: [feature]})
    #         related_filters &= feature_filters
    #     except Exception:
    #         pass

    # Apply all filters
    qs = qs.filter(related_filters)

    # Order by relevance
    # - Featured products first
    # - Higher rating
    # - Higher popularity
    qs = qs.order_by('-featured', '-rating', '-popularity')

    # Paginate the results
    return paginate_queryset(request, qs, ProductListingOutSchema, page, page_size)


# Read Single ProductListing (Retrieve)
@router.get("/product-listings/{product_listing_id}/", response=ProductListingOneOutSchema)
def retrieve_product_listing(request, product_listing_id: int):
    product_listing = get_object_or_404(ProductListing, id=product_listing_id)
    return product_listing

@router.get("/product-listings/slug/{product_listing_slug}/", response=ProductListingOneOutSchema)
@cache_response()
def retrieve_product_listing_slug(request, product_listing_slug: str):
    product_listing = get_object_or_404(ProductListing, slug=product_listing_slug)
    return ProductListingOneOutSchema.from_orm(product_listing)

# # Update ProductListing
# @router.put("/product-listings/{product_listing_id}/", response=ProductListingOutSchema)
# def update_product_listing(request, product_listing_id: int, payload: ProductListingUpdateSchema):
#     product_listing = get_object_or_404(ProductListing, id=product_listing_id)
#     for attr, value in payload.dict().items():
#         if value is not None:
#             setattr(product_listing, attr, value)
#     product_listing.save()
#     return product_listing

# Delete ProductListing
@router.delete("/product-listings/{product_listing_id}/")
def delete_product_listing(request, product_listing_id: int):
    product_listing = get_object_or_404(ProductListing, id=product_listing_id)
    product_listing.delete()
    return {"success": True}


############################ Return Exchange Policy ####################
@router.post("/return-exchange-policies/", response=ReturnExchangePolicySchema)
def create_return_exchange_policy(request, payload: ReturnExchangePolicyCreateSchema):
    policy = ReturnExchangePolicy.objects.create(**payload.dict())
    return policy

@router.get("/return-exchange-policies/", response=PaginatedResponseSchema)
def list_return_exchange_policies(request, page: int = 1, page_size: int = 10):
    queryset = ReturnExchangePolicy.objects.all()
    return paginate_queryset(request, queryset, ReturnExchangePolicySchema, page, page_size)

@router.get("/return-exchange-policies/{policy_id}/", response=ReturnExchangePolicySchema)
def retrieve_return_exchange_policy(request, policy_id: int):
    policy = get_object_or_404(ReturnExchangePolicy, id=policy_id)
    return policy

@router.put("/return-exchange-policies/{policy_id}/", response=ReturnExchangePolicySchema)
def update_return_exchange_policy(request, policy_id: int, payload: ReturnExchangePolicyUpdateSchema):
    policy = get_object_or_404(ReturnExchangePolicy, id=policy_id)
    for attr, value in payload.dict(exclude_unset=True).items():
        setattr(policy, attr, value)
    policy.save()
    return policy

@router.delete("/return-exchange-policies/{policy_id}/", response={204: None})
def delete_return_exchange_policy(request, policy_id: int):
    policy = get_object_or_404(ReturnExchangePolicy, id=policy_id)
    policy.delete()
    return 204, None


@router.get("/features/", response=PaginatedResponseSchema)
def features(request,  page: int = 1, page_size: int = 10, product_listing_id: int = None , ordering: str = None,):
    qs = Feature.objects.all()


    if product_listing_id:
        qs = qs.filter(product_listing_id__id=product_listing_id)

    if ordering:
        qs = qs.order_by(ordering)

    return paginate_queryset(request, qs, FeatureOutSchema, page, page_size)



@router.get("/product-listing-images/", response=PaginatedResponseSchema)
def product_listing_images(request,  page: int = 1, page_size: int = 10, product_listing_id: int = None , ordering: str = None,):
    qs = ProductListingImage.objects.all()

    if product_listing_id:
        qs = qs.filter(product_listing_id__id=product_listing_id)

    if ordering:
        qs = qs.order_by(ordering)

    return paginate_queryset(request, qs, ProductListingImageOutSchema, page, page_size)


@router.post("/product-listing-images/", response=ProductListingImageOutSchema)
def create_product_listing_image(
    request,
    product_listing_id: int = Form(...),
    image: UploadedFile = File(...),
    alt_text: str = Form(None)
):
    product_listing_image = ProductListingImage.objects.create(
        product_listing_id=product_listing_id,
        image=image,
        alt_text=alt_text
    )
    return product_listing_image

@router.delete("/product-listing-images/{image_id}/")
def delete_product_listing_image(request, image_id: int):
    from .models import ProductListingImage
    image = get_object_or_404(ProductListingImage, id=image_id)
    image.delete()
    return {"success": True}

@router.post("/variants/", response=VariantSchema)
def create_variant(request, payload: VariantCreateSchema):
    variant = Variant.objects.create(**payload.dict())
    return variant

@router.put("/variants/{variant_id}/", response=VariantSchema)
def update_variant(request, variant_id: int, payload: VariantUpdateSchema):
    variant = get_object_or_404(Variant, id=variant_id)
    for attr, value in payload.dict(exclude_unset=True).items():
        if value is not None:
            setattr(variant, attr, value)
    variant.save()
    return variant
